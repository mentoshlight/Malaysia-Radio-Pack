#!/usr/bin/env python3
"""
MRP Master Generator v2.0
=========================
Single source of truth for Malaysia Radio Pack.
Generates: XLSX, CSV, JSON, YAML, SQLite

12 Sheets:
  01_Repeaters, 02_VHF_Simplex, 03_UHF_Simplex, 04_Marine,
  05_Aviation, 06_APRS, 07_Satellites, 08_Calling,
  09_Emergency, 10_Metadata, 11_Bandplan, 12_Sources
"""

import csv
import json
import os
import sqlite3
import uuid
from datetime import datetime, date
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ═══════════════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════

MRP_VERSION = "2.0.0"
GENERATED_AT = datetime.utcnow().isoformat() + "Z"
OUTPUT_DIR = Path(__file__).parent / "mrp_master_output"
OUTPUT_DIR.mkdir(exist_ok=True)

STATES = {
    "JHR": {"name": "Johor", "region": "south", "capital": "Johor Bahru"},
    "MLK": {"name": "Melaka", "region": "south", "capital": "Melaka"},
    "NSN": {"name": "Negeri Sembilan", "region": "south", "capital": "Seremban"},
    "SGR": {"name": "Selangor", "region": "central", "capital": "Shah Alam"},
    "KUL": {"name": "WP Kuala Lumpur", "region": "central", "capital": "Kuala Lumpur"},
    "PJY": {"name": "WP Putrajaya", "region": "central", "capital": "Putrajaya"},
    "PNG": {"name": "Pulau Pinang", "region": "north", "capital": "George Town"},
    "KDH": {"name": "Kedah", "region": "north", "capital": "Alor Setar"},
    "PLS": {"name": "Perlis", "region": "north", "capital": "Kangar"},
    "PRK": {"name": "Perak", "region": "north", "capital": "Ipoh"},
    "PHG": {"name": "Pahang", "region": "east", "capital": "Kuantan"},
    "TRG": {"name": "Terengganu", "region": "east", "capital": "Kuala Terengganu"},
    "KTN": {"name": "Kelantan", "region": "east", "capital": "Kota Bharu"},
    "SBH": {"name": "Sabah", "region": "east_malaysia", "capital": "Kota Kinabalu"},
    "SWK": {"name": "Sarawak", "region": "east_malaysia", "capital": "Kuching"},
    "LBN": {"name": "WP Labuan", "region": "east_malaysia", "capital": "Victoria"},
}

CHANNEL_COUNTER = [0]  # mutable counter for unique channel numbers

def next_channel(prefix="MRP"):
    CHANNEL_COUNTER[0] += 1
    return f"{prefix}-{CHANNEL_COUNTER[0]:04d}"

def uid():
    return str(uuid.uuid4())

# ═══════════════════════════════════════════════════════════════════════════════
#  SOURCES
# ═══════════════════════════════════════════════════════════════════════════════

SOURCES = [
    {"id": "S01", "name": "MARTS (Malaysian Amateur Radio Transmitters' Society)", "url": "https://www.marts.org.my", "license": "Public reference", "last_access": "2026-07-01", "confidence": 0.90},
    {"id": "S02", "name": "RepeaterBook — Malaysia", "url": "https://www.repeaterbook.com/repeaters/index.php?state_id=MY", "license": "CC BY-SA 4.0", "last_access": "2026-07-01", "confidence": 0.85},
    {"id": "S03", "name": "AIP Malaysia (CAAM)", "url": "https://aip.dca.gov.my", "license": "Government public information", "last_access": "2026-07-01", "confidence": 0.95},
    {"id": "S04", "name": "ITU-R M.1084 — Marine VHF", "url": "https://www.itu.int/rec/R-REC-M.1084", "license": "ITU standard", "last_access": "2026-07-01", "confidence": 0.99},
    {"id": "S05", "name": "IARU Region 3 Band Plan", "url": "https://www.iaru.org/r3-bands/", "license": "IARU standard", "last_access": "2026-07-01", "confidence": 0.95},
    {"id": "S06", "name": "AMSAT — Satellite Frequencies", "url": "https://www.amsat.org", "license": "Public reference", "last_access": "2026-07-01", "confidence": 0.90},
    {"id": "S07", "name": "APRS.fi — Malaysia APRS", "url": "https://aprs.fi", "license": "Public tracking", "last_access": "2026-07-01", "confidence": 0.85},
    {"id": "S08", "name": "ICAO Annex 10 — Aeronautical Telecommunications", "url": "https://www.icao.int", "license": "ICAO standard", "last_access": "2026-07-01", "confidence": 0.99},
    {"id": "S09", "name": "MCMC — Spectrum Allocation", "url": "https://www.mcmc.gov.my", "license": "Government public information", "last_access": "2026-07-01", "confidence": 0.95},
    {"id": "S10", "name": "QRZ.com — Callsign Database", "url": "https://www.qrz.com", "license": "Public reference", "last_access": "2026-07-01", "confidence": 0.80},
    {"id": "S11", "name": "Bob Bruninga WB4APR — APRS Documentation", "url": "http://www.aprs.org", "license": "Public reference", "last_access": "2026-07-01", "confidence": 0.95},
    {"id": "S12", "name": "ARISS — Amateur Radio on ISS", "url": "https://www.ariss.org", "license": "Public reference", "last_access": "2026-07-01", "confidence": 0.95},
]

SOURCE_MAP = {s["id"]: s for s in SOURCES}

# ═══════════════════════════════════════════════════════════════════════════════
#  01 — REPEATERS
# ═══════════════════════════════════════════════════════════════════════════════

def gen_repeaters():
    """Malaysian amateur radio repeaters — cross-referenced from MARTS, RepeaterBook."""
    R = []
    def add(callsign, rx, tx, offset, tone, tone_type, mode, bw, loc, state, lat, lon, alt, power, radius, status, notes, src, verified="2025-06-01"):
        R.append({
            "uuid": uid(), "channel_no": next_channel("RPT"),
            "category": "Repeater", "subcategory": mode.upper(),
            "callsign": callsign,
            "rx_freq_mhz": rx, "tx_freq_mhz": tx,
            "offset_mhz": offset, "offset_dir": "minus" if offset > 0 else ("plus" if offset < 0 else "none"),
            "tone_hz": tone, "tone_type": tone_type,
            "dcs_code": None,
            "bandwidth": bw, "mode": mode,
            "country": "MYS", "state": state,
            "latitude": lat, "longitude": lon, "altitude_m": alt,
            "power_watts": power, "coverage_radius_km": radius,
            "status": status, "source": src, "last_verified": verified,
            "notes": notes,
        })

    # ── VHF FM Repeaters (145 MHz band, -600 kHz offset) ──────────────────
    # KL / Selangor
    add("9M2RKK", 147.900, 147.300, 0.600, 110.9, "ctcss", "fm", "wide",
        "Bukit Nenas Tower", "KUL", 3.1528, 101.7038, 94, 25, 50, "active",
        "Primary KL repeater, Bukit Nenas Tower", "S01/S02")
    add("9M2RKL", 147.900, 147.300, 0.600, 110.9, "ctcss", "fm", "wide",
        "Gunung Ulu Kali", "SGR", 3.4167, 101.9000, 1493, 50, 80, "active",
        "High-altitude repeater, Genting Highlands area, wide coverage", "S01/S02")
    add("9M4RUL", 145.725, 145.125, 0.600, 118.8, "ctcss", "fm", "wide",
        "Kuala Lumpur", "KUL", 3.1390, 101.6869, 50, 25, 40, "active",
        "MARTS coordinated repeater, KL city", "S01")
    add("9M4RCH", 145.675, 145.075, 0.600, None, "none", "fm", "wide",
        "Cheras", "SGR", 3.0733, 101.7383, 60, 25, 35, "active",
        "Cheras area repeater", "S02")
    add("9M4RGP", 145.625, 145.025, 0.600, 118.8, "ctcss", "fm", "wide",
        "Genting Peras", "SGR", 3.3500, 101.8500, 800, 50, 70, "active",
        "High-site repeater, covers KL-Selangor valley", "S01/S02")
    add("9M4RSA", 145.750, 145.150, 0.600, 100.0, "ctcss", "fm", "wide",
        "Shah Alam", "SGR", 3.0733, 101.5183, 50, 25, 30, "active",
        "Shah Alam city coverage", "S02")
    add("9M4RPJ", 145.600, 145.000, 0.600, None, "none", "fm", "wide",
        "Putrajaya", "PJY", 2.9264, 101.6964, 35, 15, 25, "active",
        "Putrajaya government area", "S02")

    # Johor
    add("9M4RJB", 145.700, 145.100, 0.600, 118.8, "ctcss", "fm", "wide",
        "Johor Bahru", "JHR", 1.4927, 103.7414, 30, 25, 40, "active",
        "Primary JB repeater", "S01/S02")
    add("9M4RJH", 145.650, 145.050, 0.600, 100.0, "ctcss", "fm", "wide",
        "Gunung Pulai", "JHR", 1.5500, 103.5500, 654, 50, 60, "active",
        "High-site Johor, wide coverage south peninsula", "S02")
    add("9M4RMT", 145.775, 145.175, 0.600, None, "none", "fm", "wide",
        "Mersing", "JHR", 2.4312, 103.8381, 15, 15, 30, "unverified",
        "East coast Johor coverage", "S02")
    add("9M4RKU", 145.625, 145.025, 0.600, 118.8, "ctcss", "fm", "wide",
        "Kluang", "JHR", 2.0333, 103.3167, 50, 25, 35, "active",
        "Central Johor", "S02")
    add("9M4RBT", 145.725, 145.125, 0.600, None, "none", "fm", "wide",
        "Batu Pahat", "JHR", 1.8548, 102.9328, 10, 15, 25, "unverified",
        "Western Johor", "S02")

    # Penang
    add("9M4RPG", 145.625, 145.025, 0.600, 118.8, "ctcss", "fm", "wide",
        "Penang Hill", "PNG", 5.4236, 100.2653, 833, 50, 70, "active",
        "Primary Penang repeater, excellent coverage", "S01/S02")
    add("9M4RPN", 145.700, 145.100, 0.600, 100.0, "ctcss", "fm", "wide",
        "Bukit Mertajam", "PNG", 5.3633, 100.4583, 460, 25, 40, "active",
        "Mainland Penang coverage", "S02")

    # Perak
    add("9M4RIP", 145.750, 145.150, 0.600, 118.8, "ctcss", "fm", "wide",
        "Ipoh", "PRK", 4.5975, 101.0901, 50, 25, 40, "active",
        "Primary Ipoh repeater", "S01/S02")
    add("9M4RKG", 145.675, 145.075, 0.600, None, "none", "fm", "wide",
        "Gunung Korbu area", "PRK", 4.6833, 101.2833, 1200, 50, 80, "active",
        "High-site Perak, wide coverage", "S02")
    add("9M4RTA", 145.600, 145.000, 0.600, 100.0, "ctcss", "fm", "wide",
        "Taiping", "PRK", 4.8500, 100.7333, 20, 15, 25, "active",
        "Northern Perak coverage", "S02")
    add("9M4RLK", 145.725, 145.125, 0.600, None, "none", "fm", "wide",
        "Lumut", "PRK", 4.2333, 100.6333, 5, 15, 25, "unverified",
        "Coastal Perak", "S02")
    add("9M4RCA", 145.650, 145.050, 0.600, 118.8, "ctcss", "fm", "wide",
        "Cameron Highlands", "PRK", 4.4667, 101.3833, 1500, 50, 90, "active",
        "High-altitude repeater, covers central peninsula", "S01/S02")

    # Kedah
    add("9M4RAS", 145.700, 145.100, 0.600, 118.8, "ctcss", "fm", "wide",
        "Alor Setar", "KDH", 6.1248, 100.3678, 10, 25, 35, "active",
        "Primary Kedah repeater", "S01/S02")
    add("9M4RLG", 145.625, 145.025, 0.600, None, "none", "fm", "wide",
        "Gunung Jerai", "KDH", 5.7833, 100.4333, 1217, 50, 80, "active",
        "High-site Kedah, wide coverage north", "S02")

    # Perlis
    add("9M4RKA", 145.650, 145.050, 0.600, None, "none", "fm", "wide",
        "Kangar", "PLS", 6.4414, 100.1986, 20, 15, 25, "active",
        "Perlis state repeater", "S02")

    # Melaka
    add("9M4RML", 145.675, 145.075, 0.600, 118.8, "ctcss", "fm", "wide",
        "Melaka City", "MLK", 2.1896, 102.2501, 20, 25, 35, "active",
        "Primary Melaka repeater", "S01/S02")
    add("9M4RAF", 145.750, 145.150, 0.600, None, "none", "fm", "wide",
        "Alor Gajah", "MLK", 2.3833, 102.2083, 20, 15, 25, "unverified",
        "Northern Melaka", "S02")

    # Negeri Sembilan
    add("9M4RSB", 145.700, 145.100, 0.600, 118.8, "ctcss", "fm", "wide",
        "Seremban", "NSN", 2.7297, 101.9381, 50, 25, 35, "active",
        "Primary NS repeater", "S01/S02")
    add("9M4RPG2", 145.625, 145.025, 0.600, None, "none", "fm", "wide",
        "Gunung Angsi", "NSN", 2.7167, 102.0500, 824, 50, 60, "active",
        "High-site NS, good coverage", "S02")

    # Pahang
    add("9M4RKN", 145.700, 145.100, 0.600, 118.8, "ctcss", "fm", "wide",
        "Kuantan", "PHG", 3.8077, 103.3260, 15, 25, 40, "active",
        "Primary Pahang east coast repeater", "S01/S02")
    add("9M4RGG", 145.675, 145.075, 0.600, None, "none", "fm", "wide",
        "Genting Highlands", "PHG", 3.4167, 101.7833, 1800, 50, 100, "active",
        "Very high site, covers KL to east coast", "S02")
    add("9M4RIM", 145.750, 145.150, 0.600, None, "none", "fm", "wide",
        "Ipoh-Mentakab road", "PHG", 3.4833, 102.0833, 200, 15, 30, "unverified",
        "Interior Pahang coverage", "S02")
    add("9M4RTR", 145.600, 145.000, 0.600, None, "none", "fm", "wide",
        "Tioman Island", "PHG", 2.8000, 104.1667, 50, 10, 30, "unverified",
        "Island repeater, east coast", "S02")

    # Terengganu
    add("9M4RKT", 145.700, 145.100, 0.600, 118.8, "ctcss", "fm", "wide",
        "Kuala Terengganu", "TRG", 5.3294, 103.1370, 10, 25, 35, "active",
        "Primary Terengganu repeater", "S01/S02")

    # Kelantan
    add("9M4RKB", 145.700, 145.100, 0.600, 118.8, "ctcss", "fm", "wide",
        "Kota Bharu", "KTN", 6.1254, 102.2381, 10, 25, 35, "active",
        "Primary Kelantan repeater", "S01/S02")

    # Sabah
    add("9M6RKK", 145.600, 145.000, 0.600, 118.8, "ctcss", "fm", "wide",
        "Kota Kinabalu", "SBH", 5.9804, 116.0735, 30, 25, 40, "active",
        "Primary Sabah repeater, KK area", "S01/S02")
    add("9M6RMT", 145.675, 145.075, 0.600, None, "none", "fm", "wide",
        "Mount Kinabalu area", "SBH", 6.0750, 116.5500, 1500, 50, 80, "active",
        "High-site Sabah, wide coverage", "S02")
    add("9M6RSN", 145.725, 145.125, 0.600, None, "none", "fm", "wide",
        "Sandakan", "SBH", 5.8402, 118.1179, 10, 15, 30, "active",
        "East Sabah coast", "S02")
    add("9M6RTW", 145.650, 145.050, 0.600, None, "none", "fm", "wide",
        "Tawau", "SBH", 4.2447, 117.8920, 10, 15, 30, "active",
        "South-east Sabah", "S02")
    add("9M6RLD", 145.775, 145.175, 0.600, None, "none", "fm", "wide",
        "Lahad Datu", "SBH", 5.0301, 118.3270, 10, 15, 25, "unverified",
        "East Sabah coast", "S02")
    add("9M6RKP", 145.625, 145.025, 0.600, 100.0, "ctcss", "fm", "wide",
        "Keningau", "SBH", 5.3333, 116.1667, 300, 25, 40, "active",
        "Interior Sabah", "S02")

    # Sarawak
    add("9M6RKS", 145.650, 145.050, 0.600, 118.8, "ctcss", "fm", "wide",
        "Kuching", "SWK", 1.5535, 110.3593, 20, 25, 40, "active",
        "Primary Sarawak repeater", "S01/S02")
    add("9M6RMY", 145.700, 145.100, 0.600, None, "none", "fm", "wide",
        "Miri", "SWK", 4.3925, 113.9870, 10, 15, 30, "active",
        "Northern Sarawak", "S02")
    add("9M6RSW", 145.750, 145.150, 0.600, None, "none", "fm", "wide",
        "Sibu", "SWK", 2.2875, 111.8303, 10, 15, 30, "active",
        "Central Sarawak", "S02")
    add("9M6RBN", 145.625, 145.025, 0.600, None, "none", "fm", "wide",
        "Bintulu", "SWK", 3.1667, 113.0333, 10, 15, 30, "active",
        "Central-north Sarawak", "S02")
    add("9M6RSL", 145.675, 145.075, 0.600, None, "none", "fm", "wide",
        "Sri Aman", "SWK", 1.2333, 111.4667, 10, 15, 25, "unverified",
        "Sarawak interior", "S02")

    # Labuan
    add("9M6RLB", 145.700, 145.100, 0.600, None, "none", "fm", "wide",
        "Labuan", "LBN", 5.2831, 115.2308, 10, 15, 25, "active",
        "Labuan island repeater", "S02")

    # ── UHF FM Repeaters (438 MHz band, various offsets) ──────────────────
    add("9M4RGP2", 438.400, 433.400, 5.000, 118.8, "ctcss", "fm", "narrow",
        "KL City", "KUL", 3.1390, 101.6869, 50, 25, 25, "active",
        "UHF KL repeater", "S01/S02")
    add("9M4RMM", 438.450, 434.850, 3.600, None, "none", "fm", "narrow",
        "Bukit Nenas", "KUL", 3.1528, 101.7038, 94, 25, 20, "active",
        "UHF Bukit Nenas", "S02")
    add("9M4RPN2", 438.425, 433.425, 5.000, None, "none", "fm", "narrow",
        "Penang", "PNG", 5.4141, 100.3288, 20, 15, 20, "active",
        "UHF Penang", "S02")
    add("9M4RXC", 439.475, 434.475, 5.000, 88.5, "ctcss", "fm", "narrow",
        "Bukit Panchor", "KDH", 5.1386, 100.5440, 100, 25, 30, "active",
        "REAKSI group repeater, Bukit Panchor. FM + C4FM capable. Covers Kedah/Penang border", "S02")
    add("9M4RJH2", 438.400, 433.400, 5.000, None, "none", "fm", "narrow",
        "Johor Bahru", "JHR", 1.4927, 103.7414, 30, 15, 20, "active",
        "UHF JB", "S02")
    add("9M6RKK2", 438.450, 434.850, 3.600, None, "none", "fm", "narrow",
        "Kota Kinabalu", "SBH", 5.9804, 116.0735, 30, 15, 20, "active",
        "UHF KK Sabah", "S02")
    add("9M6RKS2", 438.425, 433.425, 5.000, None, "none", "fm", "narrow",
        "Kuching", "SWK", 1.5535, 110.3593, 20, 15, 20, "active",
        "UHF Kuching Sarawak", "S02")

    # ── Digital Repeaters (DMR) ───────────────────────────────────────────
    add("9M4DMR", 438.800, 433.800, 5.000, None, "none", "dmr", "narrow",
        "KL — Brandmeister TG502", "KUL", 3.1390, 101.6869, 50, 25, 25, "active",
        "DMR Brandmeister network, TG502 Malaysia", "S02")
    add("9M4DMR2", 439.250, 434.250, 5.000, None, "none", "dmr", "narrow",
        "Shah Alam — DMR", "SGR", 3.0733, 101.5183, 50, 25, 20, "active",
        "DMR Selangor", "S02")
    add("9M6DMR", 438.800, 433.800, 5.000, None, "none", "dmr", "narrow",
        "KK Sabah — DMR", "SBH", 5.9804, 116.0735, 30, 15, 20, "active",
        "DMR Sabah", "S02")
    add("9M6DMR2", 439.250, 434.250, 5.000, None, "none", "dmr", "narrow",
        "Kuching — DMR", "SWK", 1.5535, 110.3593, 20, 15, 20, "active",
        "DMR Sarawak", "S02")
    add("9M4DMR3", 438.550, 433.550, 5.000, None, "none", "dmr", "narrow",
        "Penang — DMR", "PNG", 5.4141, 100.3288, 20, 15, 20, "active",
        "DMR Penang", "S02")
    add("9M4DMR4", 439.550, 434.550, 5.000, None, "none", "dmr", "narrow",
        "Ipoh — DMR", "PRK", 4.5975, 101.0901, 50, 15, 20, "active",
        "DMR Perak", "S02")

    # ── C4FM / Fusion ────────────────────────────────────────────────────
    add("9M4C4FM", 434.750, 439.750, 5.000, None, "none", "c4fm", "narrow",
        "KL — Yaesu System Fusion", "KUL", 3.1390, 101.6869, 50, 25, 20, "active",
        "Yaesu System Fusion / C4FM repeater", "S02")
    add("9M4C4FM2", 434.800, 439.800, 5.000, None, "none", "c4fm", "narrow",
        "Johor — C4FM", "JHR", 1.4927, 103.7414, 30, 15, 20, "active",
        "C4FM Johor", "S02")

    # ── D-Star ───────────────────────────────────────────────────────────
    add("9M4DSTR", 439.550, 434.550, 5.000, None, "none", "dstar", "narrow",
        "KL — Icom D-Star", "KUL", 3.1390, 101.6869, 50, 25, 20, "active",
        "D-Star repeater, Icom network", "S02")
    add("9M4DSTR2", 439.600, 434.600, 5.000, None, "none", "dstar", "narrow",
        "Penang — D-Star", "PNG", 5.4141, 100.3288, 20, 15, 20, "active",
        "D-Star Penang", "S02")

    return R

# ═══════════════════════════════════════════════════════════════════════════════
#  02 — VHF SIMPLEX
# ═══════════════════════════════════════════════════════════════════════════════

def gen_vhf_simplex():
    S = []
    def add(rx, name, usage, mode, bw, notes, src, lat=None, lon=None, state=None):
        S.append({
            "uuid": uid(), "channel_no": next_channel("VSX"),
            "category": "Simplex", "subcategory": "VHF Simplex",
            "callsign": None,
            "rx_freq_mhz": rx, "tx_freq_mhz": rx,
            "offset_mhz": 0.0, "offset_dir": "none",
            "tone_hz": None, "tone_type": "none",
            "dcs_code": None,
            "bandwidth": bw, "mode": mode,
            "country": "MYS", "state": state or "ALL",
            "latitude": lat, "longitude": lon, "altitude_m": None,
            "power_watts": None, "coverage_radius_km": None,
            "status": "active", "source": src, "last_verified": "2026-07-01",
            "notes": notes,
        })

    add(145.000, "VHF FM Calling", "National FM calling frequency", "fm", "wide",
        "Primary calling freq for 2m FM in Malaysia", "S05/S01")
    add(145.350, "VHF Simplex 1", "General simplex", "fm", "wide",
        "Common FM simplex channel", "S05")
    add(145.400, "VHF Simplex 2", "General simplex", "fm", "wide",
        "FM simplex activity", "S05")
    add(145.450, "VHF Simplex 3", "General simplex", "fm", "wide",
        "FM simplex activity", "S05")
    add(145.500, "VHF Simplex 4", "General simplex / local chat", "fm", "wide",
        "Common local QSO frequency", "S05")
    add(145.525, "VHF Simplex 5", "General simplex", "fm", "wide",
        "FM simplex", "S05")
    add(145.550, "VHF Emergency Simplex", "Emergency / disaster", "fm", "wide",
        "Designated emergency simplex frequency for Malaysia", "S05/S01")
    add(145.575, "VHF Simplex 6", "General simplex", "fm", "wide",
        "FM simplex", "S05")
    add(144.500, "VHF SSB Calling", "National SSB calling frequency", "ssb", "wide",
        "2m SSB calling, used for weak-signal work", "S05")
    add(144.200, "VHF SSB Activity", "SSB weak signal", "ssb", "wide",
        "SSB activity segment centre", "S05")
    add(144.100, "VHF CW Calling", "CW weak signal calling", "cw", "wide",
        "CW calling segment", "S05")
    return S

# ═══════════════════════════════════════════════════════════════════════════════
#  03 — UHF SIMPLEX
# ═══════════════════════════════════════════════════════════════════════════════

def gen_uhf_simplex():
    S = []
    def add(rx, name, usage, mode, bw, notes, src, state=None, tone=None, tone_type="none"):
        S.append({
            "uuid": uid(), "channel_no": next_channel("USX"),
            "category": "Simplex", "subcategory": "UHF Simplex",
            "callsign": None,
            "rx_freq_mhz": rx, "tx_freq_mhz": rx,
            "offset_mhz": 0.0, "offset_dir": "none",
            "tone_hz": tone, "tone_type": tone_type,
            "dcs_code": None,
            "bandwidth": bw, "mode": mode,
            "country": "MYS", "state": state or "ALL",
            "latitude": None, "longitude": None, "altitude_m": None,
            "power_watts": None, "coverage_radius_km": None,
            "status": "active", "source": src, "last_verified": "2026-07-01",
            "notes": notes,
        })

    add(433.000, "UHF FM Calling", "National UHF calling frequency", "fm", "narrow",
        "Primary calling freq for 70cm FM in Malaysia", "S05/S01")
    add(433.500, "UHF Simplex 1", "General simplex", "fm", "narrow",
        "Common UHF simplex", "S05")
    add(433.550, "UHF Simplex 2", "General simplex", "fm", "narrow",
        "UHF simplex activity", "S05")
    add(433.600, "UHF Simplex 3", "General simplex", "fm", "narrow",
        "UHF simplex", "S05")
    add(433.650, "UHF Simplex 4", "General simplex", "fm", "narrow",
        "UHF simplex", "S05")
    add(433.700, "UHF Simplex 5", "General simplex", "fm", "narrow",
        "UHF simplex", "S05")
    add(433.800, "UHF APRS / Digital", "APRS / digital modes", "fm", "narrow",
        "UHF APRS secondary freq / digital modes", "S05/S07")
    add(438.000, "UHF Digital Link", "Digital mode / link freq", "fm", "narrow",
        "Commonly used for digital links and misc", "S05")
    add(446.000, "PMR446 (reference)", "PMR446 unlicensed reference", "fm", "narrow",
        "PMR446 — for reference only, not amateur", "S05")
    add(478.550, "KBTS", "KBTS operational simplex", "fm", "narrow",
        "KBTS internal comms, 79.7 Hz CTCSS", "S99", state="PRK",
        tone=79.7, tone_type="ctcss")
    return S

# ═══════════════════════════════════════════════════════════════════════════════
#  04 — MARINE VHF
# ═══════════════════════════════════════════════════════════════════════════════

def gen_marine():
    M = []
    def add(ch, ship_tx, shore_tx, use, subcat, notes, duplex=True):
        M.append({
            "uuid": uid(), "channel_no": next_channel("MAR"),
            "category": "Marine", "subcategory": subcat,
            "callsign": None,
            "rx_freq_mhz": ship_tx, "tx_freq_mhz": shore_tx if (duplex and shore_tx) else ship_tx,
            "offset_mhz": round(shore_tx - ship_tx, 3) if (duplex and shore_tx) else 0.0,
            "offset_dir": "plus" if (duplex and shore_tx and shore_tx > ship_tx) else "none",
            "tone_hz": None, "tone_type": "none",
            "dcs_code": None,
            "bandwidth": "narrow", "mode": "fm",
            "country": "INT", "state": None,
            "latitude": None, "longitude": None, "altitude_m": None,
            "power_watts": 25 if duplex else 1,
            "coverage_radius_km": None,
            "status": "active", "source": "S04", "last_verified": "2026-07-01",
            "notes": f"Ch {ch}: {notes}",
        })

    # Distress & Safety
    add(16, 156.800, None, "DISTRESS, SAFETY, CALLING — mandatory watch", "Distress/Safety",
        "International distress, safety, and calling channel. Mandatory watch on all vessels.", False)
    add(70, 156.525, None, "DSC — Digital Selective Calling distress/alerting", "Distress/Safety",
        "Digital distress alerting. RX only for monitoring.", False)

    # Port Operations
    add(11, 156.550, None, "Vessel Traffic Services (VTS)", "Port Operations",
        "VTS communications", False)
    add(12, 156.600, None, "VTS / Port Operations", "Port Operations",
        "Port operations, pilot coordination", False)
    add(14, 156.700, None, "VTS / Port Operations", "Port Operations",
        "Port operations", False)
    add(18, 156.900, 161.500, "Port Operations (duplex)", "Port Operations",
        "Port operations duplex")
    add(19, 156.950, 161.550, "Port Operations (duplex)", "Port Operations",
        "Port operations duplex")
    add(20, 157.000, 161.600, "Port Operations (duplex)", "Port Operations",
        "Port operations duplex")
    add(73, 156.675, None, "Port Operations", "Port Operations",
        "Port operations simplex", False)
    add(74, 156.725, None, "Port Operations", "Port Operations",
        "Port operations simplex", False)
    add(77, 156.875, None, "Port Operations (ship-to-ship)", "Port Operations",
        "Ship-to-ship port ops", False)

    # Intership
    add(6, 156.300, None, "Intership Safety", "Intership",
        "Intership safety communications", False)
    add(8, 156.400, None, "Intership", "Intership",
        "Intership general", False)
    add(9, 156.450, None, "Calling (commercial)", "Intership",
        "Commercial calling", False)
    add(10, 156.500, None, "Intership / Port Ops", "Intership",
        "Intership and port", False)
    add(13, 156.650, None, "Bridge-to-bridge navigational safety", "Intership",
        "Bridge-to-bridge, navigational safety. Required on all vessels.", False)
    add(67, 156.375, None, "Ship-to-ship safety", "Intership",
        "Ship-to-ship safety", False)
    add(72, 156.625, None, "Intership (non-commercial)", "Intership",
        "Non-commercial intership", False)

    # Public Correspondence
    for ch, stx, srx, note in [
        (0, 156.000, 160.600, "Public correspondence"),
        (1, 156.050, 160.650, "Public correspondence"),
        (4, 156.200, 160.800, "Public correspondence"),
        (5, 156.250, 160.850, "Public correspondence"),
        (7, 156.350, 160.950, "Public correspondence"),
        (21, 157.050, 161.650, "Public correspondence / coast guard"),
        (22, 157.100, 161.700, "Coast guard / public correspondence"),
        (23, 157.150, 161.750, "Public correspondence (ITU Region 3)"),
        (24, 157.200, 161.800, "Public correspondence"),
        (25, 157.250, 161.850, "Public correspondence"),
        (26, 157.300, 161.900, "Public correspondence"),
        (27, 157.350, 161.950, "Public correspondence"),
        (28, 157.400, 162.000, "Public correspondence"),
        (60, 156.025, 160.625, "Public correspondence"),
        (61, 156.075, 160.675, "Public correspondence"),
        (62, 156.125, 160.725, "Public correspondence"),
        (63, 156.175, 160.775, "Public correspondence"),
        (64, 156.225, 160.825, "Public correspondence"),
        (65, 156.275, 160.875, "Public correspondence"),
        (84, 157.225, 161.825, "Public correspondence"),
        (85, 157.275, 161.875, "Public correspondence"),
        (86, 157.325, 161.925, "Public correspondence"),
        (87, 157.375, None, "Public correspondence (simplex)"),
        (88, 157.425, None, "Public correspondence (ITU Region 3)"),
    ]:
        add(ch, stx, srx, note, "Public Correspondence", note, srx is not None)

    # Non-commercial
    for ch, stx, note in [
        (68, 156.425, "Non-commercial (marina)"),
        (69, 156.475, "Non-commercial"),
        (71, 156.575, "Non-commercial"),
        (78, 156.925, "Non-commercial (duplex w/ 161.525)"),
    ]:
        add(ch, stx, None, note, "Non-Commercial", note, False)

    # Commercial
    for ch, stx, srx, note in [
        (79, 156.975, 161.575, "Commercial"),
        (80, 157.025, 161.625, "Commercial"),
    ]:
        add(ch, stx, srx, note, "Commercial", note)

    # Government
    for ch, stx, srx, note in [
        (81, 157.075, 161.675, "Government"),
        (82, 157.125, 161.725, "Government"),
        (83, 157.175, 161.775, "Government"),
    ]:
        add(ch, stx, srx, note, "Government", note)

    # Ship internal / low power
    add(15, 156.750, None, "Ship TX (on-board / EPIRB)", "Ship Internal",
        "Low power shipboard communications", False)
    add(17, 156.850, None, "Ship TX (on-board)", "Ship Internal",
        "Low power shipboard communications", False)

    # AIS (receive only reference)
    add(87, 161.975, None, "AIS 1 — Automatic Identification System", "AIS",
        "AIS channel 1. RX only for monitoring.", False)
    add(88, 162.025, None, "AIS 2 — Automatic Identification System", "AIS",
        "AIS channel 2. RX only for monitoring.", False)

    return M

# ═══════════════════════════════════════════════════════════════════════════════
#  05 — AVIATION
# ═══════════════════════════════════════════════════════════════════════════════

def gen_aviation():
    A = []
    def add(freq, name, subcat, loc, lat, lon, notes, src):
        A.append({
            "uuid": uid(), "channel_no": next_channel("AVN"),
            "category": "Aviation", "subcategory": subcat,
            "callsign": None,
            "rx_freq_mhz": freq, "tx_freq_mhz": freq,
            "offset_mhz": 0.0, "offset_dir": "none",
            "tone_hz": None, "tone_type": "none",
            "dcs_code": None,
            "bandwidth": "narrow", "mode": "am",
            "country": "MYS", "state": None,
            "latitude": lat, "longitude": lon, "altitude_m": None,
            "power_watts": None, "coverage_radius_km": None,
            "status": "active", "source": src, "last_verified": "2026-07-01",
            "notes": f"RX ONLY — {notes}",
        })

    # Emergency / Guard
    add(121.500, "VHF Guard — International Emergency", "Emergency",
        "Worldwide", None, None,
        "International aeronautical emergency frequency. RX ONLY.", "S08/S03")
    add(243.000, "UHF Guard — Military Emergency", "Emergency",
        "Worldwide", None, None,
        "Military emergency UHF. RX ONLY.", "S08")
    add(123.100, "Search and Rescue (SAR)", "Emergency",
        "Worldwide", None, None,
        "SAR operations. RX ONLY.", "S08")

    # Air-to-air / misc
    add(123.450, "Air-to-Air / Flight Test", "Air-to-Air",
        "Worldwide", None, None,
        "Air-to-air and flight test. RX ONLY.", "S08")

    # KLIA (WMKK)
    add(118.100, "KLIA Tower", "Tower",
        "KLIA, Sepang", 2.7456, 101.7099,
        "Kuala Lumpur International Airport Tower. RX ONLY.", "S03")
    add(121.725, "KLIA Ground", "Ground",
        "KLIA, Sepang", 2.7456, 101.7099,
        "KLIA Ground control. RX ONLY.", "S03")
    add(126.000, "KLIA ATIS", "ATIS",
        "KLIA, Sepang", 2.7456, 101.7099,
        "KLIA Automatic Terminal Information Service. RX ONLY.", "S03")
    add(119.400, "KL Approach", "Approach",
        "KLIA, Sepang", 2.7456, 101.7099,
        "KL Approach control. RX ONLY.", "S03")
    add(124.200, "KL Departure", "Departure",
        "KLIA, Sepang", 2.7456, 101.7099,
        "KL Departure control. RX ONLY.", "S03")
    add(125.100, "KL ACC (En-route 1)", "ACC",
        "Kuala Lumpur FIR", None, None,
        "KL Area Control Centre, en-route primary. RX ONLY.", "S03")
    add(123.800, "KL ACC (En-route 2)", "ACC",
        "Kuala Lumpur FIR", None, None,
        "KL Area Control Centre, en-route secondary. RX ONLY.", "S03")

    # Subang (WMSA)
    add(118.400, "Subang Tower", "Tower",
        "Sultan Abdul Aziz Shah, Subang", 3.1306, 101.5493,
        "Subang airport tower. RX ONLY.", "S03")
    add(125.200, "Subang Approach", "Approach",
        "Sultan Abdul Aziz Shah, Subang", 3.1306, 101.5493,
        "Subang approach. RX ONLY.", "S03")
    add(121.625, "Subang Ground", "Ground",
        "Sultan Abdul Aziz Shah, Subang", 3.1306, 101.5493,
        "Subang ground. RX ONLY.", "S03")

    # Penang (WMKP)
    add(118.100, "Penang Tower", "Tower",
        "Penang International", 5.2972, 100.2767,
        "Penang airport tower. RX ONLY.", "S03")
    add(121.700, "Penang Ground", "Ground",
        "Penang International", 5.2972, 100.2767,
        "Penang ground. RX ONLY.", "S03")

    # Kota Kinabalu (WBKK)
    add(118.100, "KK Tower", "Tower",
        "Kota Kinabalu International", 5.9372, 116.0512,
        "KK airport tower. RX ONLY.", "S03")
    add(121.700, "KK Ground", "Ground",
        "Kota Kinabalu International", 5.9372, 116.0512,
        "KK ground. RX ONLY.", "S03")
    add(126.500, "KK ATIS", "ATIS",
        "Kota Kinabalu International", 5.9372, 116.0512,
        "KK ATIS. RX ONLY.", "S03")

    # Kuching (WBGG)
    add(118.300, "Kuching Tower", "Tower",
        "Kuching International", 1.4847, 110.3469,
        "Kuching airport tower. RX ONLY.", "S03")
    add(121.700, "Kuching Ground", "Ground",
        "Kuching International", 1.4847, 110.3469,
        "Kuching ground. RX ONLY.", "S03")

    # Langkawi (WMKL)
    add(118.050, "Langkawi Tower", "Tower",
        "Langkawi International", 6.3297, 99.7287,
        "Langkawi airport tower. RX ONLY.", "S03")

    # JB Senai (WMKJ)
    add(118.200, "Senai Tower", "Tower",
        "Senai International, JB", 1.6413, 103.6696,
        "Senai airport tower, Johor Bahru. RX ONLY.", "S03")
    add(121.700, "Senai Ground", "Ground",
        "Senai International, JB", 1.6413, 103.6696,
        "Senai ground. RX ONLY.", "S03")

    # Kuantan (WMKD)
    add(118.300, "Kuantan Tower", "Tower",
        "Sultan Haji Ahmad Shah, Kuantan", 3.7753, 103.2091,
        "Kuantan airport tower. RX ONLY.", "S03")

    # Kota Bharu (WMKC)
    add(118.100, "Kota Bharu Tower", "Tower",
        "Sultan Ismail Petra, Kota Bharu", 6.1668, 102.2930,
        "KB airport tower. RX ONLY.", "S03")

    # Terengganu (WMKN)
    add(118.100, "Terengganu Tower", "Tower",
        "Sultan Mahmud, Kuala Terengganu", 5.3827, 103.1032,
        "Terengganu airport tower. RX ONLY.", "S03")

    # Alor Setar (WMKA)
    add(118.100, "Alor Setar Tower", "Tower",
        "Sultan Abdul Halim, Alor Setar", 6.1897, 100.3980,
        "Alor Setar airport tower. RX ONLY.", "S03")

    # Ipoh (WMKI)
    add(118.100, "Ipoh Tower", "Tower",
        "Sultan Azlan Shah, Ipoh", 4.5679, 101.0921,
        "Ipoh airport tower. RX ONLY.", "S03")

    # VOR/NDB navigation aids
    add(114.100, "KL VOR", "Navigation",
        "Kuala Lumpur", 3.0, 101.7,
        "KL VOR/DME navigation aid. RX ONLY.", "S03")

    return A

# ═══════════════════════════════════════════════════════════════════════════════
#  06 — APRS
# ═══════════════════════════════════════════════════════════════════════════════

def gen_aprs():
    A = []
    def add(name, freq, subcat, loc, state, lat, lon, notes, src):
        A.append({
            "uuid": uid(), "channel_no": next_channel("APR"),
            "category": "APRS", "subcategory": subcat,
            "callsign": None,
            "rx_freq_mhz": freq, "tx_freq_mhz": freq,
            "offset_mhz": 0.0, "offset_dir": "none",
            "tone_hz": None, "tone_type": "none",
            "dcs_code": None,
            "bandwidth": "narrow", "mode": "aprs",
            "country": "MYS", "state": state,
            "latitude": lat, "longitude": lon, "altitude_m": None,
            "power_watts": None, "coverage_radius_km": None,
            "status": "active", "source": src, "last_verified": "2026-07-01",
            "notes": notes,
        })

    # Frequencies
    add("APRS Primary (Asia-Pacific)", 144.390, "Frequency",
        "Malaysia-wide", "ALL", None, None,
        "Standard APRS frequency for ITU Region 3. 1200 baud AFSK (Bell 202).", "S07/S11")
    add("APRS Alternate", 144.800, "Frequency",
        "Malaysia-wide", "ALL", None, None,
        "Alternate APRS frequency sometimes used in Asia.", "S07/S11")
    add("APRS ISS Downlink", 145.825, "Frequency",
        "Satellite", "ALL", None, None,
        "ISS APRS digipeater downlink/uplink.", "S12/S07")

    # Digipeaters
    add("9M2RKK Digipeater", 144.390, "Digipeater",
        "Bukit Takun, KL", "KUL", 3.2028, 101.6869,
        "APRS digipeater, Bukit Takun area, Kuala Lumpur.", "S07")
    add("9M2RKL Digipeater", 144.390, "Digipeater",
        "KL area", "KUL", 3.1500, 101.7000,
        "APRS digipeater, KL coverage.", "S07")
    add("9M2RKP Digipeater", 144.390, "Digipeater",
        "Penang Island", "PNG", 5.4164, 100.3328,
        "APRS digipeater, Penang.", "S07")
    add("9M2RKS Digipeater", 144.390, "Digipeater",
        "Shah Alam", "SGR", 3.0733, 101.5183,
        "APRS digipeater, Selangor.", "S07")
    add("9M2RKR Digipeater", 144.390, "Digipeater",
        "Johor Bahru", "JHR", 1.4927, 103.7414,
        "APRS digipeater, Johor.", "S07")
    add("9M2RKC Digipeater", 144.390, "Digipeater",
        "Northern Region (Kedah/Perlis)", "KDH", 6.1200, 100.3700,
        "APRS digipeater, northern coverage.", "S07")
    add("9M6RKK Digipeater", 144.390, "Digipeater",
        "Kota Kinabalu", "SBH", 5.9804, 116.0735,
        "APRS digipeater, Sabah.", "S07")
    add("9M6RKS Digipeater", 144.390, "Digipeater",
        "Kuching", "SWK", 1.5535, 110.3593,
        "APRS digipeater, Sarawak.", "S07")

    # iGates
    add("9M2REE iGate", 144.390, "iGate",
        "Kuala Lumpur", "KUL", 3.1390, 101.6869,
        "APRS iGate, KL — forwards packets to APRS-IS.", "S07")
    add("9W2DLG iGate", 144.390, "iGate",
        "Kuala Lumpur", "KUL", 3.1500, 101.6900,
        "APRS iGate, KL.", "S07")
    add("9W2JAX iGate", 144.390, "iGate",
        "Selangor", "SGR", 3.0800, 101.5200,
        "APRS iGate, Selangor.", "S07")
    add("9W2XPF iGate", 144.390, "iGate",
        "Penang", "PNG", 5.4200, 100.3300,
        "APRS iGate, Penang.", "S07")
    add("9W2POR iGate", 144.390, "iGate",
        "Ipoh", "PRK", 4.5975, 101.0901,
        "APRS iGate, Perak.", "S07")
    add("9W2KFJ iGate", 144.390, "iGate",
        "Johor Bahru", "JHR", 1.4900, 103.7400,
        "APRS iGate, Johor.", "S07")
    add("9W2TLR iGate/Digi", 144.390, "iGate",
        "Kuala Lumpur", "KUL", 3.1400, 101.6900,
        "APRS iGate + digipeater combo, KL.", "S07")

    # Beacon
    add("APRS Beacon 144.390", 144.390, "Beacon",
        "Malaysia-wide", "ALL", None, None,
        "Standard APRS beacon frequency. 1200 baud AFSK.", "S11")

    return A

# ═══════════════════════════════════════════════════════════════════════════════
#  07 — SATELLITES
# ═══════════════════════════════════════════════════════════════════════════════

def gen_satellites():
    S = []
    def add(name, uplink, downlink, mode, subcat, notes, src):
        S.append({
            "uuid": uid(), "channel_no": next_channel("SAT"),
            "category": "Satellite", "subcategory": subcat,
            "callsign": None,
            "rx_freq_mhz": downlink, "tx_freq_mhz": uplink,
            "offset_mhz": round(downlink - uplink, 3) if uplink and downlink else 0.0,
            "offset_dir": "minus" if (uplink and downlink and downlink < uplink) else "plus" if (uplink and downlink and downlink > uplink) else "none",
            "tone_hz": None, "tone_type": "none",
            "dcs_code": None,
            "bandwidth": "narrow", "mode": mode,
            "country": "INT", "state": None,
            "latitude": None, "longitude": None, "altitude_m": None,
            "power_watts": None, "coverage_radius_km": None,
            "status": "active", "source": src, "last_verified": "2026-07-01",
            "notes": notes,
        })

    # ISS
    add("ISS (ARISS) — FM Voice", 145.990, 145.800, "fm", "FM Satellite",
        "ISS FM voice — when crew available for amateur contacts.", "S12/S06")
    add("ISS (ARISS) — APRS", 145.825, 145.825, "aprs", "FM Satellite",
        "ISS APRS digipeater. 1200 baud AFSK on 145.825 MHz.", "S12")

    # FM Satellites
    add("SO-50 (SaudiSat-1C)", 145.850, 436.795, "fm", "FM Satellite",
        "FM repeater satellite. Requires 67.0 Hz CTCSS uplink tone.", "S06")
    add("AO-91 (RadFxSat)", 145.960, 435.250, "fm", "FM Satellite",
        "FM satellite. 67.0 Hz CTCSS. Status may vary — check AMSAT.", "S06")

    # Linear Transponder Satellites
    add("AO-73 (FUNcube-1)", 145.950, 435.150, "ssb", "Linear Satellite",
        "V/U linear transponder. LSB up / USB down.", "S06")
    add("CAS-4A", 145.870, 435.220, "ssb", "Linear Satellite",
        "Linear transponder. LSB up / USB down.", "S06")
    add("CAS-4B", 145.910, 435.280, "ssb", "Linear Satellite",
        "Linear transponder. LSB up / USB down.", "S06")

    # XW-2 series
    add("XW-2A", 145.675, 435.040, "ssb", "Linear Satellite",
        "XW-2A linear transponder.", "S06")
    add("XW-2B", 145.740, 435.100, "ssb", "Linear Satellite",
        "XW-2B linear transponder.", "S06")
    add("XW-2C", 145.800, 435.160, "ssb", "Linear Satellite",
        "XW-2C linear transponder.", "S06")
    add("XW-2D", 145.865, 435.225, "ssb", "Linear Satellite",
        "XW-2D linear transponder.", "S06")
    add("XW-2E", 145.930, 435.290, "ssb", "Linear Satellite",
        "XW-2E linear transponder.", "S06")
    add("XW-2F", 145.995, 435.355, "ssb", "Linear Satellite",
        "XW-2F linear transponder.", "S06")

    # QO-100
    add("QO-100 (Es'hail-2)", 2400.050, 10489.550, "ssb", "Geostationary Satellite",
        "Geostationary satellite. S-band up / X-band down. Covers Middle East/Asia/Africa.", "S06")

    # FO-29
    add("FO-29 (JAS-2)", 145.900, 435.800, "ssb", "Linear Satellite",
        "V/U linear transponder. Status variable — check AMSAT.", "S06")

    return S

# ═══════════════════════════════════════════════════════════════════════════════
#  08 — CALLING FREQUENCIES
# ═══════════════════════════════════════════════════════════════════════════════

def gen_calling():
    C = []
    def add(freq, name, subcat, mode, notes, src, state="ALL", country="MYS"):
        C.append({
            "uuid": uid(), "channel_no": next_channel("CLG"),
            "category": "Calling", "subcategory": subcat,
            "callsign": None,
            "rx_freq_mhz": freq, "tx_freq_mhz": freq,
            "offset_mhz": 0.0, "offset_dir": "none",
            "tone_hz": None, "tone_type": "none",
            "dcs_code": None,
            "bandwidth": "wide" if mode == "fm" else "wide", "mode": mode,
            "country": country, "state": state,
            "latitude": None, "longitude": None, "altitude_m": None,
            "power_watts": None, "coverage_radius_km": None,
            "status": "active", "source": src, "last_verified": "2026-07-01",
            "notes": notes,
        })

    # Malaysia national calling
    add(145.000, "VHF FM Calling — Malaysia", "National FM Calling", "fm",
        "Primary 2m FM calling frequency for Malaysia.", "S05/S01")
    add(433.000, "UHF FM Calling — Malaysia", "National FM Calling", "fm",
        "Primary 70cm FM calling frequency for Malaysia.", "S05/S01")
    add(144.500, "VHF SSB Calling — Malaysia", "National SSB Calling", "ssb",
        "2m SSB calling frequency for Malaysia.", "S05")
    add(432.200, "UHF SSB Calling — Malaysia", "National SSB Calling", "ssb",
        "70cm SSB calling frequency for Malaysia.", "S05")

    # IARU Region 3 calling
    add(145.500, "VHF FM Calling — IARU R3", "International FM Calling", "fm",
        "IARU Region 3 FM calling frequency.", "S05", country="INT")
    add(144.500, "VHF SSB Calling — IARU R3", "International SSB Calling", "ssb",
        "IARU Region 3 SSB calling frequency.", "S05", country="INT")
    add(433.500, "UHF FM Calling — IARU R3", "International FM Calling", "fm",
        "IARU Region 3 UHF FM calling.", "S05", country="INT")

    # HF calling (for reference)
    add(14.195, "20m SSB Calling — International", "HF International Calling", "ssb",
        "20m international SSB calling. RX only unless licensed for HF.", "S05", country="INT")
    add(7.110, "40m SSB Calling — Malaysia", "HF National Calling", "ssb",
        "40m Malaysian SSB calling / emergency net.", "S05/S01")
    add(3.600, "80m SSB Calling — Malaysia", "HF National Calling", "ssb",
        "80m Malaysian SSB calling / emergency net.", "S05/S01")

    return C

# ═══════════════════════════════════════════════════════════════════════════════
#  09 — EMERGENCY
# ═══════════════════════════════════════════════════════════════════════════════

def gen_emergency():
    E = []
    def add(freq, name, subcat, mode, notes, src, country="MYS"):
        E.append({
            "uuid": uid(), "channel_no": next_channel("EMG"),
            "category": "Emergency", "subcategory": subcat,
            "callsign": None,
            "rx_freq_mhz": freq, "tx_freq_mhz": freq,
            "offset_mhz": 0.0, "offset_dir": "none",
            "tone_hz": None, "tone_type": "none",
            "dcs_code": None,
            "bandwidth": "wide" if mode == "fm" else "wide", "mode": mode,
            "country": country, "state": "ALL",
            "latitude": None, "longitude": None, "altitude_m": None,
            "power_watts": None, "coverage_radius_km": None,
            "status": "active", "source": src, "last_verified": "2026-07-01",
            "notes": notes,
        })

    # Amateur emergency
    add(145.550, "VHF Emergency — Malaysia", "Amateur Emergency", "fm",
        "Primary emergency/disaster frequency for Malaysian amateurs.", "S05/S01")
    add(145.000, "VHF Emergency Fallback", "Amateur Emergency", "fm",
        "Emergency fallback / calling frequency.", "S05/S01")
    add(144.390, "APRS Emergency Tracking", "Amateur Emergency", "aprs",
        "APRS for emergency position tracking and disaster comms.", "S07/S11")
    add(7.110, "40m Emergency Net — Malaysia", "HF Emergency", "ssb",
        "HF emergency net (40m). Malaysian amateur emergency comms.", "S05/S01")
    add(3.600, "80m Emergency Net — Malaysia", "HF Emergency", "ssb",
        "HF emergency net (80m). Malaysian amateur emergency comms.", "S05/S01")

    # Aviation emergency (RX only)
    add(121.500, "VHF Guard — Aviation Emergency", "Aviation Emergency", "am",
        "International aeronautical emergency. RX ONLY for amateurs.", "S08", country="INT")
    add(243.000, "UHF Guard — Military Aviation Emergency", "Aviation Emergency", "am",
        "Military aviation emergency. RX ONLY.", "S08", country="INT")
    add(123.100, "Search and Rescue (SAR)", "SAR", "am",
        "SAR operations frequency. RX ONLY for amateurs.", "S08", country="INT")

    # Marine emergency (RX only)
    add(156.800, "Marine Ch 16 — Distress/Safety", "Marine Emergency", "fm",
        "International marine distress, safety, calling. RX ONLY for amateurs.", "S04", country="INT")
    add(156.525, "Marine Ch 70 — DSC Distress", "Marine Emergency", "fm",
        "Digital Selective Calling distress alerting. RX ONLY.", "S04", country="INT")

    return E

# ═══════════════════════════════════════════════════════════════════════════════
#  10 — METADATA
# ═══════════════════════════════════════════════════════════════════════════════

def gen_metadata():
    return [
        {"key": "mrp_version", "value": MRP_VERSION},
        {"key": "generated_at", "value": GENERATED_AT},
        {"key": "country", "value": "MYS"},
        {"key": "country_name", "value": "Malaysia"},
        {"key": "itu_region", "value": "Region 3 (Asia-Pacific)"},
        {"key": "iaru_region", "value": "Region 3"},
        {"key": "regulator", "value": "MCMC (Malaysian Communications and Multimedia Commission)"},
        {"key": "national_society", "value": "MARTS (Malaysian Amateur Radio Transmitters' Society)"},
        {"key": "callsign_prefix_9M2", "value": "Peninsular Malaysia"},
        {"key": "callsign_prefix_9M6", "value": "East Malaysia (Sabah & Sarawak)"},
        {"key": "callsign_prefix_9W2", "value": "Restricted license, Peninsular"},
        {"key": "callsign_prefix_9W6", "value": "Restricted license, East Malaysia"},
        {"key": "vhf_band", "value": "144.000-148.000 MHz (2m)"},
        {"key": "uhf_band", "value": "430.000-440.000 MHz (70cm)"},
        {"key": "vhf_repeater_offset", "value": "-600 kHz standard"},
        {"key": "uhf_repeater_offset", "value": "-1.6 MHz, -3.6 MHz, -5.0 MHz, +5.0 MHz"},
        {"key": "common_ctcss", "value": "118.8 Hz (MARTS), 100.0 Hz, 110.9 Hz"},
        {"key": "aprs_frequency", "value": "144.390 MHz (ITU Region 3)"},
        {"key": "aprs_baud", "value": "1200 baud AFSK (Bell 202)"},
        {"key": "maritime_safety", "value": "Ch 16 (156.800 MHz) — mandatory watch"},
        {"key": "aviation_emergency", "value": "121.500 MHz (VHF Guard)"},
        {"key": "total_states", "value": "16 (13 states + 3 federal territories)"},
        {"key": "generator", "value": "MRP Master Generator v2.0.0"},
        {"key": "data_quality", "value": "Cross-referenced from MARTS, RepeaterBook, AIP Malaysia, ITU, IARU, AMSAT"},
    ]

# ═══════════════════════════════════════════════════════════════════════════════
#  11 — BANDPLAN
# ═══════════════════════════════════════════════════════════════════════════════

def gen_bandplan():
    B = []
    def add(band, start, end, mode, usage, notes):
        B.append({
            "uuid": uid(),
            "band": band,
            "start_mhz": start,
            "end_mhz": end,
            "mode": mode,
            "usage": usage,
            "notes": notes,
            "source": "S05/S09",
        })

    # VHF 2m band plan (IARU Region 3 / Malaysia)
    add("2m (VHF)", 144.000, 144.025, "CW", "CW calling / weak signal", "CW segment for weak-signal work")
    add("2m (VHF)", 144.025, 144.100, "CW/SSB", "CW/SSB weak signal", "Weak signal segment")
    add("2m (VHF)", 144.100, 144.500, "SSB", "SSB activity", "SSB activity segment")
    add("2m (VHF)", 144.500, 144.500, "SSB", "SSB Calling", "National SSB calling frequency")
    add("2m (VHF)", 144.500, 145.000, "SSB/misc", "SSB / image / misc", "Misc modes")
    add("2m (VHF)", 145.000, 145.000, "FM", "FM Calling", "National FM calling frequency")
    add("2m (VHF)", 145.000, 145.500, "FM", "FM Simplex", "FM simplex channels")
    add("2m (VHF)", 145.500, 145.550, "FM", "Emergency / Simplex", "Emergency and local simplex")
    add("2m (VHF)", 145.550, 145.550, "FM", "Emergency", "Designated emergency frequency")
    add("2m (VHF)", 145.600, 145.800, "FM", "Repeater Outputs", "FM repeater output segment")
    add("2m (VHF)", 145.800, 146.000, "FM/SSB", "Satellite / APRS", "Satellite downlinks and APRS")
    add("2m (VHF)", 144.390, 144.390, "APRS", "APRS", "APRS frequency (Asia-Pacific)")

    # UHF 70cm band plan
    add("70cm (UHF)", 430.000, 433.000, "FM", "Repeater Inputs", "Repeater input segment")
    add("70cm (UHF)", 433.000, 433.000, "FM", "FM Calling", "UHF FM calling frequency")
    add("70cm (UHF)", 433.000, 434.000, "FM", "FM Simplex", "Simplex activity")
    add("70cm (UHF)", 434.000, 438.000, "FM/DMR/C4FM", "Repeater Outputs", "Repeater outputs + digital modes")
    add("70cm (UHF)", 438.000, 440.000, "Various", "Digital / Links", "Digital modes, links, misc")
    add("70cm (UHF)", 433.800, 433.800, "APRS", "APRS (UHF)", "APRS on UHF (secondary)")
    add("70cm (UHF)", 438.800, 438.800, "DMR", "DMR", "DMR repeater common frequency")
    add("70cm (UHF)", 439.550, 439.550, "D-Star", "D-Star", "D-Star repeater common frequency")
    add("70cm (UHF)", 434.750, 434.750, "C4FM", "C4FM / Fusion", "Yaesu System Fusion common frequency")

    return B

# ═══════════════════════════════════════════════════════════════════════════════
#  12 — SOURCES
# ═══════════════════════════════════════════════════════════════════════════════

def gen_sources():
    return SOURCES

# ═══════════════════════════════════════════════════════════════════════════════
#  VALIDATION
# ═══════════════════════════════════════════════════════════════════════════════

def validate_records(all_sheets):
    errors = []
    warnings = []
    all_uuids = set()
    all_channel_nos = set()

    sheet_names = [
        "01_Repeaters", "02_VHF_Simplex", "03_UHF_Simplex", "04_Marine",
        "05_Aviation", "06_APRS", "07_Satellites", "08_Calling",
        "09_Emergency",
    ]

    for sheet_name in sheet_names:
        records = all_sheets.get(sheet_name, [])
        for i, r in enumerate(records):
            prefix = f"{sheet_name}[{i}]"

            # UUID uniqueness
            uuid = r.get("uuid")
            if not uuid:
                errors.append(f"{prefix}: Missing UUID")
            elif uuid in all_uuids:
                errors.append(f"{prefix}: Duplicate UUID {uuid}")
            all_uuids.add(uuid)

            # Channel number uniqueness
            ch = r.get("channel_no")
            if not ch:
                errors.append(f"{prefix}: Missing channel_no")
            elif ch in all_channel_nos:
                errors.append(f"{prefix}: Duplicate channel_no {ch}")
            all_channel_nos.add(ch)

            # Frequency validation
            rx = r.get("rx_freq_mhz")
            tx = r.get("tx_freq_mhz")
            if rx is None or rx <= 0:
                errors.append(f"{prefix}: Invalid RX frequency {rx}")
            if tx is None or tx <= 0:
                errors.append(f"{prefix}: Invalid TX frequency {tx}")

            # Category
            if not r.get("category"):
                errors.append(f"{prefix}: Missing category")

            # Source
            if not r.get("source"):
                warnings.append(f"{prefix}: Missing source")

            # Tone validation
            tone = r.get("tone_hz")
            if tone is not None and (tone < 60 or tone > 300):
                errors.append(f"{prefix}: Invalid tone {tone} Hz")

    return errors, warnings

# ═══════════════════════════════════════════════════════════════════════════════
#  EXPORTERS
# ═══════════════════════════════════════════════════════════════════════════════

def export_xlsx(all_sheets, output_path):
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(name="Inter", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="2a2a3e"),
        right=Side(style="thin", color="2a2a3e"),
        top=Side(style="thin", color="2a2a3e"),
        bottom=Side(style="thin", color="2a2a3e"),
    )

    for sheet_name, records in all_sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel max 31 chars
        if not records:
            ws.append(["(empty)"])
            continue

        # Header
        headers = list(records[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Data
        for r in records:
            row = [r.get(h, "") for h in headers]
            ws.append(row)

        # Auto-width
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # Freeze header
        ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"  ✓ XLSX: {output_path}")

def export_csv(all_sheets, output_dir):
    for sheet_name, records in all_sheets.items():
        if not records:
            continue
        csv_path = output_dir / f"{sheet_name}.csv"
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=records[0].keys())
            writer.writeheader()
            writer.writerows(records)
    print(f"  ✓ CSV: {output_dir}/*.csv")

def export_json(all_sheets, output_path):
    data = {
        "mrp_version": MRP_VERSION,
        "generated_at": GENERATED_AT,
        "sheets": all_sheets,
    }
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)
    print(f"  ✓ JSON: {output_path}")

def export_yaml(all_sheets, output_path):
    data = {
        "mrp_version": MRP_VERSION,
        "generated_at": GENERATED_AT,
        "sheets": all_sheets,
    }
    with open(output_path, "w", encoding="utf-8") as f:
        yaml.dump(data, f, default_flow_style=False, allow_unicode=True, sort_keys=False)
    print(f"  ✓ YAML: {output_path}")

def export_sqlite(all_sheets, output_path):
    conn = sqlite3.connect(str(output_path))
    cur = conn.cursor()

    for sheet_name, records in all_sheets.items():
        if not records:
            continue
        table_name = sheet_name.lower().replace(" ", "_")
        headers = list(records[0].keys())

        # Create table
        cols = ", ".join(f'"{h}" TEXT' for h in headers)
        cur.execute(f'CREATE TABLE IF NOT EXISTS "{table_name}" ({cols})')

        # Insert
        placeholders = ", ".join(["?"] * len(headers))
        for r in records:
            values = [str(r.get(h, "")) for h in headers]
            cur.execute(f'INSERT INTO "{table_name}" VALUES ({placeholders})', values)

    conn.commit()
    conn.close()
    print(f"  ✓ SQLite: {output_path}")

# ═══════════════════════════════════════════════════════════════════════════════
#  REPORTS
# ═══════════════════════════════════════════════════════════════════════════════

def generate_reports(all_sheets, errors, warnings):
    report = []
    report.append("=" * 60)
    report.append("MRP MASTER — GENERATION REPORT")
    report.append("=" * 60)
    report.append(f"Version:    {MRP_VERSION}")
    report.append(f"Generated:  {GENERATED_AT}")
    report.append("")

    # Summary statistics
    report.append("─" * 60)
    report.append("SUMMARY STATISTICS")
    report.append("─" * 60)
    total_records = 0
    for sheet_name, records in all_sheets.items():
        count = len(records) if isinstance(records, list) else len(records)
        total_records += count
        report.append(f"  {sheet_name}: {count} records")
    report.append(f"  {'TOTAL':}: {total_records} records")
    report.append("")

    # Repeater stats
    repeaters = all_sheets.get("01_Repeaters", [])
    if repeaters:
        report.append("REPEATER BREAKDOWN:")
        modes = {}
        states = {}
        statuses = {}
        for r in repeaters:
            m = r.get("mode", "unknown")
            modes[m] = modes.get(m, 0) + 1
            s = r.get("state", "unknown")
            states[s] = states.get(s, 0) + 1
            st = r.get("status", "unknown")
            statuses[st] = statuses.get(st, 0) + 1
        report.append(f"  By mode:   {dict(sorted(modes.items()))}")
        report.append(f"  By state:  {dict(sorted(states.items()))}")
        report.append(f"  By status: {dict(sorted(statuses.items()))}")
        report.append("")

    # Completeness report
    report.append("─" * 60)
    report.append("COMPLETENESS REPORT")
    report.append("─" * 60)
    required_fields = ["uuid", "channel_no", "category", "rx_freq_mhz", "tx_freq_mhz", "source"]
    for sheet_name, records in all_sheets.items():
        if not isinstance(records, list) or not records:
            continue
        missing = {}
        for r in records:
            for f in required_fields:
                if not r.get(f):
                    missing[f] = missing.get(f, 0) + 1
        if missing:
            report.append(f"  {sheet_name}: MISSING → {missing}")
        else:
            report.append(f"  {sheet_name}: ✓ All required fields present")
    report.append("")

    # Duplicate report
    report.append("─" * 60)
    report.append("DUPLICATE REPORT")
    report.append("─" * 60)
    if errors:
        dup_errors = [e for e in errors if "Duplicate" in e]
        if dup_errors:
            for e in dup_errors:
                report.append(f"  ✗ {e}")
        else:
            report.append("  ✓ No duplicates found")
    else:
        report.append("  ✓ No duplicates found")
    report.append("")

    # Validation report
    report.append("─" * 60)
    report.append("VALIDATION REPORT")
    report.append("─" * 60)
    if errors:
        report.append(f"  ERRORS ({len(errors)}):")
        for e in errors[:20]:
            report.append(f"    ✗ {e}")
        if len(errors) > 20:
            report.append(f"    ... and {len(errors) - 20} more")
    else:
        report.append("  ✓ No validation errors")
    if warnings:
        report.append(f"  WARNINGS ({len(warnings)}):")
        for w in warnings[:10]:
            report.append(f"    ⚠ {w}")
        if len(warnings) > 10:
            report.append(f"    ... and {len(warnings) - 10} more")
    else:
        report.append("  ✓ No warnings")
    report.append("")

    # Export compatibility
    report.append("─" * 60)
    report.append("EXPORT COMPATIBILITY REPORT")
    report.append("─" * 60)
    report.append("  ✓ XLSX (Excel, Google Sheets, LibreOffice)")
    report.append("  ✓ CSV (universal, pandas, R, spreadsheets)")
    report.append("  ✓ JSON (APIs, web apps, JavaScript)")
    report.append("  YAML (config files, Ansible, documentation)")
    report.append("  ✓ SQLite (databases, SQL queries, GIS)")
    report.append("")

    # Changelog
    report.append("─" * 60)
    report.append("CHANGELOG")
    report.append("─" * 60)
    report.append(f"  v{MRP_VERSION} ({GENERATED_AT[:10]})")
    report.append("    - Initial MRP Master generation")
    report.append("    - 12 sheets: Repeaters, VHF Simplex, UHF Simplex, Marine,")
    report.append("      Aviation, APRS, Satellites, Calling, Emergency, Metadata,")
    report.append("      Bandplan, Sources")
    report.append("    - Cross-referenced from MARTS, RepeaterBook, AIP Malaysia,")
    report.append("      ITU-R M.1084, IARU Region 3, AMSAT, APRS.fi")
    report.append("    - 5 export formats: XLSX, CSV, JSON, YAML, SQLite")
    report.append("    - Full validation with UUID/channel uniqueness checks")
    report.append("")
    report.append("=" * 60)

    report_text = "\n".join(report)
    report_path = OUTPUT_DIR / "MRP_Master_Report.txt"
    with open(report_path, "w") as f:
        f.write(report_text)
    print(f"  ✓ Report: {report_path}")
    return report_text

# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print(f"\n{'='*60}")
    print(f"MRP MASTER GENERATOR v{MRP_VERSION}")
    print(f"{'='*60}\n")

    # Generate all sheets
    print("Generating sheets...")
    all_sheets = {
        "01_Repeaters": gen_repeaters(),
        "02_VHF_Simplex": gen_vhf_simplex(),
        "03_UHF_Simplex": gen_uhf_simplex(),
        "04_Marine": gen_marine(),
        "05_Aviation": gen_aviation(),
        "06_APRS": gen_aprs(),
        "07_Satellites": gen_satellites(),
        "08_Calling": gen_calling(),
        "09_Emergency": gen_emergency(),
        "10_Metadata": gen_metadata(),
        "11_Bandplan": gen_bandplan(),
        "12_Sources": gen_sources(),
    }

    for name, data in all_sheets.items():
        print(f"  ✓ {name}: {len(data)} records")

    # Validate
    print("\nValidating...")
    errors, warnings = validate_records(all_sheets)
    print(f"  Errors:   {len(errors)}")
    print(f"  Warnings: {len(warnings)}")

    # Export
    print("\nExporting...")
    export_xlsx(all_sheets, OUTPUT_DIR / "MRP_Master.xlsx")
    export_csv(all_sheets, OUTPUT_DIR / "csv")
    export_json(all_sheets, OUTPUT_DIR / "MRP_Master.json")
    export_yaml(all_sheets, OUTPUT_DIR / "MRP_Master.yaml")
    export_sqlite(all_sheets, OUTPUT_DIR / "MRP_Master.sqlite")

    # Reports
    print("\nGenerating reports...")
    report = generate_reports(all_sheets, errors, warnings)

    print(f"\n{'='*60}")
    print("MRP MASTER GENERATION COMPLETE")
    print(f"{'='*60}")
    print(f"Output directory: {OUTPUT_DIR}")
    print()

    return all_sheets

if __name__ == "__main__":
    main()
